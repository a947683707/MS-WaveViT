import torch.utils.data as data
from PIL import Image
import os
import os.path
import scipy.io
import numpy as np
import csv
import cv2  # 添加这行导入
from openpyxl import load_workbook


class LIVEFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        refpath = os.path.join(root, 'refimgs')
        refname = getFileName(refpath, '.bmp')

        jp2kroot = os.path.join(root, 'jp2k')
        jp2kname = self.getDistortionTypeFileName(jp2kroot, 227)

        jpegroot = os.path.join(root, 'jpeg')
        jpegname = self.getDistortionTypeFileName(jpegroot, 233)

        wnroot = os.path.join(root, 'wn')
        wnname = self.getDistortionTypeFileName(wnroot, 174)

        gblurroot = os.path.join(root, 'gblur')
        gblurname = self.getDistortionTypeFileName(gblurroot, 174)

        fastfadingroot = os.path.join(root, 'fastfading')
        fastfadingname = self.getDistortionTypeFileName(fastfadingroot, 174)

        imgpath = jp2kname + jpegname + wnname + gblurname + fastfadingname

        dmos = scipy.io.loadmat(os.path.join(root, 'dmos_realigned.mat'))
        labels = dmos['dmos_new'].astype(np.float32)

        orgs = dmos['orgs']
        refnames_all = scipy.io.loadmat(os.path.join(root, 'refnames_all.mat'))
        refnames_all = refnames_all['refnames_all']

        sample = []

        for i in range(0, len(index)):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = train_sel * ~orgs.astype(np.bool_)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[1].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append((imgpath[item], labels[0][item]))
                # print(self.imgpath[item])
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        if self.transform is not None:
            sample = self.transform(sample)

        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

    def getDistortionTypeFileName(self, path, num):
        filename = []
        index = 1
        for i in range(0, num):
            name = '%s%s%s' % ('img', str(index), '.bmp')
            filename.append(os.path.join(path, name))
            index = index + 1
        return filename


class LIVEChallengeFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        imgpath = scipy.io.loadmat(os.path.join(root, 'Data', 'AllImages_release.mat'))
        imgpath = imgpath['AllImages_release']
        imgpath = imgpath[7:1169]
        mos = scipy.io.loadmat(os.path.join(root, 'Data', 'AllMOS_release.mat'))
        labels = mos['AllMOS_release'].astype(np.float32)
        labels = labels[0][7:1169]

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, 'Images', imgpath[item][0][0]), labels[item]))

        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length

class ESPL_LIVE_HDRFolder(data.Dataset):
    def __init__(self, root, label_file, index, transform, patch_num):
        """
        root: 图片文件夹路径
        label_file: 标签txt文件路径，格式示例：
            V_Sequoia_Remains_WardHistAdjTMO.PNG    54.7726036359    10.5247238303
        index: 你想读取的样本索引列表（整数索引，用于按文件排序索引图片）
        transform: 图片变换
        patch_num: 每张图片增强的次数
        """
        self.samples = []
        self.transform = transform

        # 读标签文件，存字典 {图片名: MOS}
        self.label_dict = {}
        with open(label_file, 'r') as f:
            for line in f:
                parts = line.strip().split()
                if len(parts) >= 2:
                    img_name = parts[0]
                    mos = float(parts[1])
                    self.label_dict[img_name] = mos

        # 读取root目录下所有图片文件名（排序）
        all_imgs = sorted(os.listdir(root))

        # 根据index选图，组合成samples (path, mos)，每张重复patch_num次
        for i in index:
            img_name = all_imgs[i]
            if img_name in self.label_dict:
                mos = self.label_dict[img_name]
                for _ in range(patch_num):
                    self.samples.append((os.path.join(root, img_name), mos))
            else:
                print(f"Warning: {img_name} not found in label file!")

    def __getitem__(self, index):
        path, target = self.samples[index]
        sample = pil_loader(path)
        if self.transform is not None:
            sample = self.transform(sample)
        return sample, target

    def __len__(self):
        return len(self.samples)



class CSIQFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        refpath = os.path.join(root, 'src_imgs')
        refname = getFileName(refpath,'.png')
        txtpath = os.path.join(root, 'csiq_label.txt')
        fh = open(txtpath, 'r')
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            imgnames.append((words[0]))
            target.append(words[1])
            ref_temp = words[0].split(".")
            refnames_all.append(ref_temp[0] + '.' + ref_temp[-1])

        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []

        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append((os.path.join(root, 'dst_imgs_all', imgnames[item]), labels[item]))
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)

        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class Koniq_10kFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):
        imgname = []
        mos_all = []
        csv_file = os.path.join(root, 'koniq10k_scores_and_distributions.csv')
        with open(csv_file) as f:
            reader = csv.DictReader(f)
            for row in reader:
                imgname.append(row['image_name'])
                mos = np.array(float(row['MOS_zscore'])).astype(np.float32)
                mos_all.append(mos)

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, '1024x768', imgname[item]), mos_all[item]))

        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class BIDFolder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):

        imgname = []
        mos_all = []

        xls_file = os.path.join(root, 'DatabaseGrades.xlsx')
        workbook = load_workbook(xls_file)
        booksheet = workbook.active
        rows = booksheet.rows
        count = 1
        for row in rows:
            count += 1
            img_num = (booksheet.cell(row=count, column=1).value)
            img_name = "DatabaseImage%04d.JPG" % (img_num)
            imgname.append(img_name)
            mos = (booksheet.cell(row=count, column=2).value)
            mos = np.array(mos)
            mos = mos.astype(np.float32)
            mos_all.append(mos)
            if count == 587:
                break

        sample = []
        for i, item in enumerate(index):
            for aug in range(patch_num):
                sample.append((os.path.join(root, imgname[item]), mos_all[item]))

        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


class TID2013Folder(data.Dataset):

    def __init__(self, root, index, transform, patch_num):
        refpath = os.path.join(root, 'reference_images')
        refname = getTIDFileName(refpath,'.bmp.BMP')
        txtpath = os.path.join(root, 'mos_with_names.txt')
        fh = open(txtpath, 'r')
        imgnames = []
        target = []
        refnames_all = []
        for line in fh:
            line = line.split('\n')
            words = line[0].split()
            imgnames.append((words[1]))
            target.append(words[0])
            ref_temp = words[1].split("_")
            refnames_all.append(ref_temp[0][1:])
        labels = np.array(target).astype(np.float32)
        refnames_all = np.array(refnames_all)

        sample = []
        for i, item in enumerate(index):
            train_sel = (refname[index[i]] == refnames_all)
            train_sel = np.where(train_sel == True)
            train_sel = train_sel[0].tolist()
            for j, item in enumerate(train_sel):
                for aug in range(patch_num):
                    sample.append((os.path.join(root, 'distorted_images', imgnames[item]), labels[item]))
        self.samples = sample
        self.transform = transform

    def __getitem__(self, index):
        """
        Args:
            index (int): Index

        Returns:
            tuple: (sample, target) where target is class_index of the target class.
        """
        path, target = self.samples[index]
        sample = pil_loader(path)
        sample = self.transform(sample)
        return sample, target

    def __len__(self):
        length = len(self.samples)
        return length


def getFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if os.path.splitext(i)[1] == suffix:
            filename.append(i)
    return filename


def getTIDFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if suffix.find(os.path.splitext(i)[1]) != -1:
            filename.append(i[1:3])
    return filename


def pil_loader(path):
    # 检查文件扩展名
    if path.lower().endswith('.exr'):
        # 使用OpenCV加载.exr文件
        img = cv2.imread(path, cv2.IMREAD_COLOR | cv2.IMREAD_ANYDEPTH)
        if img is None:
            raise ValueError(f"Cannot load .exr file: {path}")
        # 转换BGR到RGB
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        # 将float32转换为uint8 (0-255范围)
        img = np.clip(img * 255, 0, 255).astype(np.uint8)
        # 转换为PIL Image
        return Image.fromarray(img)
    else:
        # 对于其他格式，使用原来的PIL方法
        with open(path, 'rb') as f:
            img = Image.open(f)
            return img.convert('RGB')


class NarwariaFolder(data.Dataset):
    def __init__(self, root, index, transform, patch_num):
        """
        root: narwaria图像文件夹路径 (upiq_dataset/images/narwaria)
        index: 要使用的样本索引列表
        transform: 图像变换
        patch_num: 每张图片的增强次数
        """
        self.samples = []
        self.transform = transform
        
        # 读取标签文件
        csv_file = os.path.join(os.path.dirname(os.path.dirname(root)), 'upiq_subjective_scores.csv')
        label_dict = {}
        
        with open(csv_file, 'r') as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) >= 12 and 'narwaria' in row[1]:  # 确保是narwaria数据
                    img_path = row[6]  # test_file列
                    jod_score = float(row[11])  # JOD列
                    label_dict[img_path] = jod_score
        
        # 收集所有图像文件
        all_images = []
        for folder_num in range(1, 11):  # 01-10文件夹
            folder_path = os.path.join(root, f"{folder_num:02d}")
            if os.path.exists(folder_path):
                for img_file in os.listdir(folder_path):
                    if img_file.endswith('.exr'):
                        relative_path = f"narwaria/{folder_num:02d}/{img_file}"
                        if relative_path in label_dict:
                            all_images.append((os.path.join(folder_path, img_file), label_dict[relative_path]))
        
        # 根据index选择样本
        for i in index:
            if i < len(all_images):
                img_path, score = all_images[i]
                for _ in range(patch_num):
                    self.samples.append((img_path, score))
    
    def __getitem__(self, index):
        path, target = self.samples[index]
        sample = pil_loader(path)
        if self.transform is not None:
            sample = self.transform(sample)
        return sample, target
    
    def __len__(self):
        return len(self.samples)


def getFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if os.path.splitext(i)[1] == suffix:
            filename.append(i)
    return filename


def getTIDFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if suffix.find(os.path.splitext(i)[1]) != -1:
            filename.append(i[1:3])
    return filename


class KorshunovFolder(data.Dataset):
    def __init__(self, root, index, transform, patch_num):
        """
        Korshunov数据集加载器
        root: korshunov数据集根目录路径
        index: 要加载的图像索引列表
        transform: 图像变换
        patch_num: 每张图像的patch数量
        """
        self.samples = []
        self.transform = transform
        
        # 读取主观分数CSV文件
        csv_file = os.path.join(os.path.dirname(os.path.dirname(root)), 'upiq_subjective_scores.csv')
        label_dict = {}
        
        with open(csv_file, 'r') as f:
            reader = csv.reader(f)
            next(reader)  # 跳过标题行
            for row in reader:
                if len(row) >= 12 and 'korshunov' in row[1]:  # 确保是korshunov数据
                    img_path = row[6]  # test_file列
                    jod_score = float(row[11])  # JOD列
                    label_dict[img_path] = jod_score
        
        # 获取所有.exr文件
        all_images = []
        for subdir in sorted(os.listdir(root)):
            subdir_path = os.path.join(root, subdir)
            if os.path.isdir(subdir_path):
                for img_file in sorted(os.listdir(subdir_path)):
                    if img_file.endswith('.exr'):
                        # 构建相对路径以匹配CSV中的路径格式
                        relative_path = f"korshunov/{subdir}/{img_file}"
                        img_full_path = os.path.join(subdir_path, img_file)
                        
                        if relative_path in label_dict:
                            all_images.append((img_full_path, label_dict[relative_path]))
                        else:
                            print(f"Warning: {relative_path} not found in label file!")
        
        # 根据index选择图像，每张图像重复patch_num次
        for i in index:
            if i < len(all_images):
                img_path, quality_score = all_images[i]
                for _ in range(patch_num):
                    self.samples.append((img_path, quality_score))
    
    def __getitem__(self, index):
        path, target = self.samples[index]
        # 加载EXR图像
        try:
            import cv2
            # 使用OpenCV加载EXR文件
            img = cv2.imread(path, cv2.IMREAD_ANYCOLOR | cv2.IMREAD_ANYDEPTH)
            if img is not None:
                # 转换为RGB格式
                img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                # 转换为PIL Image
                from PIL import Image
                img = Image.fromarray((img * 255).astype(np.uint8))
            else:
                # 如果OpenCV无法加载，创建一个默认图像
                from PIL import Image
                img = Image.new('RGB', (512, 384), color='black')
        except Exception as e:
            print(f"Error loading {path}: {e}")
            from PIL import Image
            img = Image.new('RGB', (512, 384), color='black')
        
        if self.transform is not None:
            img = self.transform(img)
        
        return img, target
    
    def __len__(self):
        return len(self.samples)


def getFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if os.path.splitext(i)[1] == suffix:
            filename.append(i)
    return filename


def getTIDFileName(path, suffix):
    filename = []
    f_list = os.listdir(path)
    for i in f_list:
        if suffix.find(os.path.splitext(i)[1]) != -1:
            filename.append(i[1:3])
    return filename


# 删除这个重复的函数定义（第468-472行）
# def pil_loader(path):
#     with open(path, 'rb') as f:
#         img = Image.open(f)
#         return img.convert('RGB')